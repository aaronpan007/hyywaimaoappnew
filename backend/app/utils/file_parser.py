"""Parse uploaded Excel/CSV/Word files into standardized lead dicts."""

import asyncio
import base64
import csv
import io
import json
import logging

logger = logging.getLogger(__name__)

# Standard field names that the parser maps to
_STANDARD_FIELDS = [
    "company_name", "website", "country", "industry", "company_role",
    "contact_name", "email", "phone", "ai_summary", "business_match",
    "outreach_suggestion", "match_score",
]

# Column name mapping: various Chinese/English headers → Lead model fields
_COLUMN_MAP = {
    # Company info
    "公司名称": "company_name",
    "company name": "company_name",
    "company": "company_name",
    "公司": "company_name",
    "company_name": "company_name",
    "网站": "website",
    "website": "website",
    "web": "website",
    "网址": "website",
    "国家": "country",
    "country": "country",
    "地区": "country",
    "region": "country",
    "行业": "industry",
    "industry": "industry",
    "公司角色": "company_role",
    "company role": "company_role",
    "company_role": "company_role",
    "角色": "company_role",
    "类型": "company_role",
    # Contact info
    "联系人": "contact_name",
    "contact name": "contact_name",
    "contact": "contact_name",
    "contact_name": "contact_name",
    "姓名": "contact_name",
    "邮箱": "email",
    "email": "email",
    "e-mail": "email",
    "mail": "email",
    "邮件": "email",
    "电话": "phone",
    "phone": "phone",
    "tel": "phone",
    "telephone": "phone",
    # AI analysis fields
    "ai摘要": "ai_summary",
    "ai_summary": "ai_summary",
    "ai summary": "ai_summary",
    "摘要": "ai_summary",
    "业务匹配": "business_match",
    "business_match": "business_match",
    "business match": "business_match",
    "匹配点": "business_match",
    "开发建议": "outreach_suggestion",
    "outreach_suggestion": "outreach_suggestion",
    "outreach suggestion": "outreach_suggestion",
    "建议": "outreach_suggestion",
    # Score
    "匹配度": "match_score",
    "match score": "match_score",
    "match_score": "match_score",
    "score": "match_score",
    "评分": "match_score",
}


def _normalize_column_name(name: str) -> str:
    """Normalize a column header for mapping."""
    return name.strip().lower()


def _map_row(row: dict, mapped_headers: dict[str, str]) -> dict:
    """Map a row dict using the column mapping. Returns a lead dict."""
    lead: dict = {}
    for col_name, field_name in mapped_headers.items():
        value = row.get(col_name, "")
        if value is None:
            value = ""
        value = str(value).strip()
        lead[field_name] = value
    return lead


def _format_extra_columns(row: dict, mapped_headers: dict[str, str]) -> str:
    """Preserve useful unmapped spreadsheet cells as a free-form lead note."""
    mapped_cols = set(mapped_headers)
    extras = []
    for col_name, value in row.items():
        header = str(col_name or "").strip()
        if not header or header in mapped_cols:
            continue
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        if len(text) > 500:
            text = text[:500].rstrip() + "..."
        extras.append(f"{header}: {text}")
    if not extras:
        return ""
    return "用户上传补充信息：\n" + "\n".join(extras[:20])


def _try_parse_match_score(value: str) -> float:
    """Try to parse match_score from various formats."""
    if not value:
        return 0.0
    # Remove percentage sign
    cleaned = value.replace("%", "").strip()
    try:
        score = float(cleaned)
        # If the original had %, normalize to 0-100
        if "%" in value and score <= 1.0:
            score = score * 100
        return min(max(score, 0), 100)
    except ValueError:
        return 0.0


async def parse_uploaded_file(filename: str, base64_data: str) -> list[dict]:
    """Parse an uploaded file (Excel/CSV/Word) into a list of lead dicts.

    Args:
        filename: Original filename with extension
        base64_data: Base64-encoded file content

    Returns:
        List of dicts with standardized lead fields
    """
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    try:
        raw_bytes = base64.b64decode(base64_data)
    except Exception as e:
        raise ValueError(f"Failed to decode base64: {e}")

    if ext in ("xlsx", "xls"):
        return await _parse_excel(raw_bytes)
    elif ext == "csv":
        return await _parse_csv(raw_bytes)
    elif ext == "docx":
        return await _parse_docx(raw_bytes)
    else:
        raise ValueError(f"Unsupported file format: .{ext}")


def _build_column_mapping(headers: list[str]) -> dict[str, str]:
    """Build a mapping from original column names to lead field names."""
    mapping = {}
    for header in headers:
        normalized = _normalize_column_name(header)
        if normalized in _COLUMN_MAP:
            mapping[header] = _COLUMN_MAP[normalized]
    return mapping


async def _smart_build_column_mapping(
    headers: list[str],
    sample_rows: list[dict],
) -> dict[str, str]:
    """Build column mapping with LLM fallback for unmatched columns.

    1. Try static mapping first
    2. If coverage < 50%, call LLM to map remaining columns
    3. LLM results only supplement (never override) static matches
    """
    static_mapping = _build_column_mapping(headers)

    # Count how many headers got mapped
    mapped_count = len(static_mapping)
    total_headers = len(headers)

    # If static mapping covers 50%+ of headers, return as-is
    if total_headers == 0 or mapped_count / total_headers >= 0.5:
        return static_mapping

    # Build prompt for LLM
    unmapped_headers = [h for h in headers if h not in static_mapping]

    # Build sample data for context (first 3 rows, truncated values)
    sample_text = ""
    for i, row in enumerate(sample_rows[:3]):
        row_data = {}
        for h in unmapped_headers:
            val = row.get(h, "")
            if val is not None:
                val_str = str(val)
                row_data[h] = val_str[:100] if len(val_str) > 100 else val_str
        if row_data:
            sample_text += f"Row {i+1}: {json.dumps(row_data, ensure_ascii=False)}\n"

    if not sample_text:
        return static_mapping

    prompt = f"""You are a data mapping assistant. I have a table with these column headers:
{json.dumps(headers, ensure_ascii=False)}

Here is sample data for the unmatched columns ({json.dumps(unmapped_headers, ensure_ascii=False)}):
{sample_text}

Standard field names: {json.dumps(_STANDARD_FIELDS)}

Please map each unmatched header to the most appropriate standard field name.
Only map columns that have a clear semantic match. Ignore columns that don't fit any standard field.

Output a JSON object: {{"original_header": "standard_field", ...}}
Only output the JSON, nothing else."""

    try:
        import os
        import replicate
        from app.config import settings

        if not settings.replicate_api_token:
            return static_mapping

        if not os.environ.get("REPLICATE_API_TOKEN"):
            os.environ["REPLICATE_API_TOKEN"] = settings.replicate_api_token

        output = await asyncio.to_thread(
            replicate.run,
            settings.replicate_model,
            input={
                "messages": [
                    {"role": "user", "content": prompt},
                ],
            },
        )

        if isinstance(output, list):
            response_text = "".join(output)
        else:
            response_text = str(output)

        # Parse JSON from response
        response_text = response_text.strip()
        # Handle markdown code blocks
        if "```" in response_text:
            import re
            code_match = re.search(r"```(?:json)?\s*\n?([\s\S]*?)\n?```", response_text)
            if code_match:
                response_text = code_match.group(1).strip()

        llm_mapping = json.loads(response_text)

        if isinstance(llm_mapping, dict):
            for header, field_name in llm_mapping.items():
                # Only add if header exists in original headers, field is valid, and not already mapped
                if (header in headers
                        and header not in static_mapping
                        and field_name in _STANDARD_FIELDS):
                    static_mapping[header] = field_name

        logger.info("LLM column mapping: mapped %d/%d unmapped headers",
                     len(llm_mapping) if isinstance(llm_mapping, dict) else 0,
                     len(unmapped_headers))

    except Exception as e:
        logger.warning("LLM column mapping failed: %s", str(e)[:200])

    return static_mapping


def _rows_to_dicts(headers: list[str], rows: list[tuple]) -> list[dict]:
    """Convert list of tuples to list of dicts using headers."""
    result = []
    for row in rows:
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = value
        result.append(row_dict)
    return result


async def _parse_excel(raw_bytes: bytes) -> list[dict]:
    """Parse an Excel file."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Empty Excel file")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty Excel file")

    headers = [str(h or "").strip() for h in rows[0]]
    data_rows = _rows_to_dicts(headers, rows[1:])

    # Use smart mapping for better coverage
    mapped_headers = await _smart_build_column_mapping(headers, data_rows)

    leads = []
    for row_dict in data_rows:
        lead = _map_row(row_dict, mapped_headers)
        if lead.get("company_name"):
            extra_note = _format_extra_columns(row_dict, mapped_headers)
            if extra_note:
                lead["user_note"] = extra_note
            if "match_score" in lead:
                lead["match_score"] = _try_parse_match_score(lead.get("match_score", ""))
            leads.append(lead)

    wb.close()
    return leads


async def _parse_csv(raw_bytes: bytes) -> list[dict]:
    """Parse a CSV file."""
    # Try different encodings
    text = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb2312", "latin-1"):
        try:
            text = raw_bytes.decode(encoding)
            break
        except (UnicodeDecodeError, LookupError):
            continue

    if text is None:
        raise ValueError("Failed to decode CSV file")

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("Empty CSV file")

    headers = list(reader.fieldnames)
    sample_rows = [row for _, row in zip(range(3), reader)]
    reader.fieldnames  # reset would need re-reading; use sample for mapping

    # Use smart mapping
    mapped_headers = await _smart_build_column_mapping(headers, sample_rows)

    # Re-read for full data
    text_io = io.StringIO(text)
    reader2 = csv.DictReader(text_io)
    leads = []
    for row in reader2:
        lead = _map_row(row, mapped_headers)
        if lead.get("company_name"):
            extra_note = _format_extra_columns(row, mapped_headers)
            if extra_note:
                lead["user_note"] = extra_note
            if "match_score" in lead:
                lead["match_score"] = _try_parse_match_score(lead.get("match_score", ""))
            leads.append(lead)

    return leads


async def _parse_docx(raw_bytes: bytes) -> list[dict]:
    """Parse a Word document, extracting tables."""
    try:
        from docx import Document
    except ImportError:
        raise ValueError("python-docx not installed. Cannot parse .docx files.")

    doc = Document(io.BytesIO(raw_bytes))

    leads = []
    for table in doc.tables:
        if len(table.rows) < 2:
            continue

        # First row as headers
        headers = [cell.text.strip() for cell in table.rows[0].cells]
        data_rows = []
        for row in table.rows[1:]:
            row_dict = {}
            for i, cell in enumerate(row.cells):
                if i < len(headers):
                    row_dict[headers[i]] = cell.text.strip()
            data_rows.append(row_dict)

        mapped_headers = await _smart_build_column_mapping(headers, data_rows)

        for row_dict in data_rows:
            lead = _map_row(row_dict, mapped_headers)
            if lead.get("company_name"):
                extra_note = _format_extra_columns(row_dict, mapped_headers)
                if extra_note:
                    lead["user_note"] = extra_note
                if "match_score" in lead:
                    lead["match_score"] = _try_parse_match_score(lead.get("match_score", ""))
                leads.append(lead)

    return leads
