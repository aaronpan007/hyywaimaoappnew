from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.models.lead import Lead
from app.models.outreach_email import OutreachEmail
from app.models.task import Task
from app.schemas.common import PaginatedResponse
from app.schemas.email import OutreachEmailResponse, UpdateLeadEmailRequest


async def get_emails(
    db: AsyncSession,
    user_id: int,
    page: int = 1,
    page_size: int = 20,
) -> PaginatedResponse[OutreachEmailResponse]:
    from app.models.task import Task
    from sqlalchemy import func

    task_ids_query = select(Task.id).where(Task.user_id == user_id)
    query = (
        select(OutreachEmail)
        .where(OutreachEmail.task_id.in_(task_ids_query))
        .order_by(OutreachEmail.created_at.desc())
    )
    count_query = (
        select(func.count())
        .select_from(OutreachEmail)
        .where(OutreachEmail.task_id.in_(task_ids_query))
    )

    total = (await db.execute(count_query)).scalar() or 0
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    emails = result.scalars().all()

    items = []
    for em in emails:
        # Get lead company name
        lead_result = await db.execute(select(Lead.company_name).where(Lead.id == em.lead_id))
        lead_company = lead_result.scalar() or ""

        items.append(
            OutreachEmailResponse(
                id=em.id,
                lead_id=em.lead_id,
                lead_company_name=lead_company,
                email_subject=em.email_subject,
                email_body=em.email_body,
                send_status=em.send_status,
                sent_at=str(em.sent_at) if em.sent_at else None,
                created_at=str(em.created_at),
            )
        )

    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=(total + page_size - 1) // page_size,
    )


async def generate_emails_stub(
    db: AsyncSession,
    user_id: int,
    lead_ids: list[int],
) -> dict:
    """Phase 1 stub — Phase 2 TODO: call email-craft Pipeline."""
    return {
        "status": "stub",
        "message": f"Would generate emails for {len(lead_ids)} leads. Phase 2: will call email-craft Pipeline.",
        "lead_ids": lead_ids,
    }


async def send_emails_stub(
    db: AsyncSession,
    user_id: int,
    email_ids: list[int],
    lead_ids: list[int] | None = None,
    delay_min: int = 60,
    delay_max: int = 120,
    daily_limit: int = 50,
    dry_run: bool = True,
    send_mode: str = "immediate",
) -> dict:
    """Start the email-blast pipeline from the emails API.

    The primary UI path uses /api/tasks/start for SSE. This JSON endpoint is
    kept for compatibility and returns the created task id.
    """
    from app.services.chat_service import start_email_blast_pipeline

    selected_lead_ids = list(lead_ids or [])
    if not selected_lead_ids and email_ids:
        task_ids_query = select(Task.id).where(Task.user_id == user_id)
        result = await db.execute(
            select(OutreachEmail.lead_id)
            .join(Lead, Lead.id == OutreachEmail.lead_id)
            .where(
                OutreachEmail.id.in_(email_ids),
                Lead.task_id.in_(task_ids_query),
            )
        )
        selected_lead_ids = [lead_id for (lead_id,) in result.all()]

    result = await start_email_blast_pipeline(
        {
            "lead_ids": selected_lead_ids,
            "delay_min": delay_min,
            "delay_max": delay_max,
            "daily_limit": daily_limit,
            "dry_run": dry_run,
            "send_mode": send_mode,
        },
        db,
        user_id,
    )
    return {"status": "started", **result}


def _decode_svix_secret(secret: str) -> bytes:
    import base64

    value = (secret or "").strip()
    if value.startswith("whsec_"):
        value = value.split("_", 1)[1]
    try:
        return base64.b64decode(value)
    except Exception:
        return (secret or "").encode("utf-8")


def verify_resend_webhook_signature(
    payload: bytes,
    headers: dict[str, str],
) -> bool:
    """Verify Resend/Svix webhook signature using the configured secret."""
    import base64
    import hashlib
    import hmac

    secret = settings.resend_webhook_secret
    if not secret:
        return False

    svix_id = headers.get("svix-id") or headers.get("Svix-Id") or ""
    svix_timestamp = headers.get("svix-timestamp") or headers.get("Svix-Timestamp") or ""
    svix_signature = headers.get("svix-signature") or headers.get("Svix-Signature") or ""
    if not svix_id or not svix_timestamp or not svix_signature:
        return False

    signed_payload = f"{svix_id}.{svix_timestamp}.".encode("utf-8") + payload
    expected = base64.b64encode(
        hmac.new(_decode_svix_secret(secret), signed_payload, hashlib.sha256).digest()
    ).decode("utf-8")

    signatures = []
    for item in svix_signature.split(" "):
        if "," in item:
            version, signature = item.split(",", 1)
            if version == "v1":
                signatures.append(signature)
        elif item.startswith("v1,"):
            signatures.append(item[3:])

    return any(hmac.compare_digest(expected, signature) for signature in signatures)


def _extract_resend_email_id(payload: dict) -> str:
    data = payload.get("data")
    if not isinstance(data, dict):
        data = {}
    candidates = [
        data.get("email_id"),
        data.get("emailId"),
        data.get("id"),
        payload.get("email_id"),
        payload.get("emailId"),
    ]
    for candidate in candidates:
        if candidate:
            return str(candidate)
    return ""


def _map_resend_event_to_status(event_type: str) -> str | None:
    mapping = {
        "email.sent": "sent",
        "email.delivered": "delivered",
        "email.delivery_delayed": "sent",
        "email.bounced": "bounced",
        "email.failed": "failed",
        "email.complained": "complained",
    }
    return mapping.get(event_type)


async def handle_resend_webhook(db: AsyncSession, payload: dict) -> dict:
    event_type = str(payload.get("type") or payload.get("event") or "")
    resend_message_id = _extract_resend_email_id(payload)
    if not event_type or not resend_message_id:
        return {"status": "ignored", "reason": "missing event type or email id"}

    mapped_status = _map_resend_event_to_status(event_type)
    if mapped_status is None:
        return {"status": "ignored", "event": event_type}

    result = await db.execute(
        select(OutreachEmail).where(OutreachEmail.resend_message_id == resend_message_id)
    )
    email = result.scalar_one_or_none()
    if email is None:
        return {"status": "ignored", "reason": "email not found", "event": event_type}

    email.last_event = event_type
    email.send_status = mapped_status
    if mapped_status in {"failed", "bounced", "complained"}:
        data = payload.get("data") if isinstance(payload.get("data"), dict) else {}
        reason = data.get("reason") or data.get("message") or event_type
        email.error_message = str(reason)[:500]
    elif mapped_status == "delivered":
        email.error_message = ""

    await db.commit()
    return {
        "status": "updated",
        "email_id": email.id,
        "send_status": mapped_status,
        "event": event_type,
    }


async def update_latest_email_for_lead(
    db: AsyncSession,
    user_id: int,
    lead_id: int,
    req: UpdateLeadEmailRequest,
) -> dict:
    """Update the displayed/latest outreach email for a lead owned by the user.

    Sent emails are preserved as history. Editing a sent email creates a new draft
    with the requested changes.
    """
    from fastapi import HTTPException

    lead_result = await db.execute(
        select(Lead).where(
            Lead.id == lead_id,
            Lead.task_id.in_(select(Task.id).where(Task.user_id == user_id)),
        )
    )
    lead = lead_result.scalar_one_or_none()
    if lead is None:
        raise HTTPException(status_code=404, detail="Lead not found")

    email_result = await db.execute(
        select(OutreachEmail)
        .where(OutreachEmail.lead_id == lead_id)
        .order_by(OutreachEmail.created_at.desc())
        .limit(1)
    )
    latest_email = email_result.scalar_one_or_none()
    if latest_email is None:
        raise HTTPException(status_code=404, detail="Email draft not found")

    subject = req.email_subject if req.email_subject is not None else latest_email.email_subject
    body = req.email_body if req.email_body is not None else latest_email.email_body

    if latest_email.send_status == "sent":
        email_record = OutreachEmail(
            lead_id=lead.id,
            task_id=latest_email.task_id,
            email_subject=subject,
            email_body=body,
            send_status="draft",
        )
        db.add(email_record)
    else:
        latest_email.email_subject = subject
        latest_email.email_body = body
        email_record = latest_email

    await db.commit()
    await db.refresh(email_record)
    return {
        "id": email_record.id,
        "lead_id": lead.id,
        "email_subject": email_record.email_subject,
        "email_body": email_record.email_body,
        "send_status": email_record.send_status,
    }


async def export_emails_xlsx(
    db: AsyncSession,
    user_id: int,
    task_id: int | None = None,
) -> bytes:
    """Export leads + outreach emails as Excel.

    14 columns: 12 lead columns + Email Subject + Email Body.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from app.models.task import Task
    from app.utils.contact import clean_contact_name

    # Get leads via outreach_emails table (not by task_id on leads)
    if task_id:
        # Step 1: Get email records for this task
        email_result = await db.execute(
            select(OutreachEmail).where(OutreachEmail.task_id == task_id)
        )
        email_records = email_result.scalars().all()

        # Step 2: Get leads by lead_id from emails
        lead_ids = list({em.lead_id for em in email_records})
        if lead_ids:
            lead_result = await db.execute(
                select(Lead).where(Lead.id.in_(lead_ids)).order_by(Lead.match_score.desc())
            )
        else:
            lead_result = await db.execute(select(Lead).where(False))
        leads = lead_result.scalars().all()

        # Step 3: Build email lookup
        emails = {em.lead_id: em for em in email_records}
    else:
        task_ids_query = select(Task.id).where(Task.user_id == user_id)
        lead_result = await db.execute(
            select(Lead).where(Lead.task_id.in_(task_ids_query)).order_by(Lead.match_score.desc())
        )
        leads = lead_result.scalars().all()

        lead_ids = [l.id for l in leads]
        if lead_ids:
            email_result = await db.execute(
                select(OutreachEmail).where(OutreachEmail.lead_id.in_(lead_ids))
            )
            emails = {em.lead_id: em for em in email_result.scalars().all()}
        else:
            emails = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Emails"

    headers = [
        "Company Name", "Website", "Country", "Industry",
        "Company Role", "Contact Name", "Email", "Phone", "Match Score",
        "AI Summary", "Business Match", "Outreach Suggestion",
        "Email Subject", "Email Body",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="EFF6FF")
    header_font = Font(bold=True, color="1F2937")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for lead in leads:
        em = emails.get(lead.id)
        ws.append([
            lead.company_name,
            lead.website,
            lead.country,
            lead.industry,
            lead.company_role,
            clean_contact_name(lead.contact_name),
            lead.email,
            lead.phone,
            lead.match_score,
            lead.ai_summary,
            lead.business_match,
            lead.outreach_suggestion,
            em.email_subject if em else "",
            em.email_body if em else "",
        ])

    ws.freeze_panes = "A2"
    widths = {
        "A": 28, "B": 28, "C": 18, "D": 22,
        "E": 16, "F": 20, "G": 28, "H": 18,
        "I": 12, "J": 52, "K": 42, "L": 42,
        "M": 36, "N": 60,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 54

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
