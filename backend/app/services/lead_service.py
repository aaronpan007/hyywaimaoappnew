import io
import logging

from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.lead import Lead
from app.models.outreach_email import OutreachEmail
from app.models.task import Task

logger = logging.getLogger(__name__)
from app.models.task import Task
from app.schemas.common import PaginatedResponse
from app.schemas.lead import LeadResponse
from app.utils.contact import clean_contact_name


async def create_leads_from_files(
    db: AsyncSession, user_id: int, files: list[dict]
) -> tuple[int, int]:
    """Create an import task, parse uploaded files, dedup & save leads.

    Args:
        db: Database session
        user_id: Current user ID
        files: List of {"filename": str, "data": str(base64)}

    Returns:
        Tuple of (import_task_id, saved_lead_count)
    """
    from app.utils.file_parser import parse_uploaded_file

    import_task = Task(
        user_id=user_id,
        type="import",
        status="completed",
        params={"file_count": len(files), "file_names": [f["filename"] for f in files]},
    )
    db.add(import_task)
    await db.commit()
    await db.refresh(import_task)

    # Get existing company names for dedup
    existing_result = await db.execute(
        select(Lead.company_name).where(
            Lead.task_id.in_(select(Task.id).where(Task.user_id == user_id))
        )
    )
    existing_names = {name.lower() for (name,) in existing_result.all()}

    saved_count = 0
    for f in files:
        try:
            parsed = await parse_uploaded_file(f["filename"], f["data"])
            for lead_data in parsed:
                company_name = lead_data.get("company_name", "")
                if company_name and company_name.lower() in existing_names:
                    continue
                lead = Lead(
                    task_id=import_task.id,
                    company_name=company_name,
                    website=lead_data.get("website", ""),
                    country=lead_data.get("country", ""),
                    industry=lead_data.get("industry", ""),
                    company_role=lead_data.get("company_role", ""),
                    contact_name=lead_data.get("contact_name", ""),
                    email=lead_data.get("email", ""),
                    phone=lead_data.get("phone", ""),
                    ai_summary=lead_data.get("ai_summary", ""),
                    business_match=lead_data.get("business_match", ""),
                    outreach_suggestion=lead_data.get("outreach_suggestion", ""),
                    match_score=lead_data.get("match_score", 0) or 0,
                )
                db.add(lead)
                existing_names.add(company_name.lower())
                saved_count += 1
        except Exception as e:
            logger.warning("Failed to parse uploaded file %s: %s", f["filename"], str(e)[:200])

    await db.commit()
    return import_task.id, saved_count


def _format_source_list(task: Task | None) -> str:
    """Return a short display name for the lead source list."""
    if task is None:
        return "未知来源"

    params = task.params or {}
    created = task.created_at.strftime("%Y-%m-%d") if task.created_at else ""

    if task.type == "customer-acquisition":
        country = params.get("country") or ""
        industry = params.get("industry") or ""
        label = " ".join(part for part in [country, industry] if part).strip()
        return f"{label or '客户搜索'} - {created}".strip(" -")

    if task.type == "email-craft":
        lead_count = params.get("lead_count", 0)
        if lead_count:
            return f"开发信撰写 ({lead_count}条) - {created}".strip(" -")
        return f"开发信名单 - {created}".strip(" -")

    if task.type == "manual":
        return f"手动输入客户 - {created}".strip(" -")

    if task.type == "import":
        return f"上传客户名单 - {created}".strip(" -")

    return f"{task.type} - {created}".strip(" -")


async def get_leads(
    db: AsyncSession,
    user_id: int,
    page: int = 1,
    page_size: int = 20,
    task_id: int | None = None,
    search: str | None = None,
    country: str | None = None,
    industry: str | None = None,
    min_score: float | None = None,
    sort_by: str = "created_at",
    sort_order: str = "desc",
) -> PaginatedResponse[LeadResponse]:
    # Only show leads from tasks belonging to this user
    from app.models.task import Task

    task_ids_query = select(Task.id).where(Task.user_id == user_id)
    query = select(Lead).where(Lead.task_id.in_(task_ids_query))
    count_query = (
        select(func.count())
        .select_from(Lead)
        .where(Lead.task_id.in_(task_ids_query))
    )

    if task_id is not None:
        query = query.where(Lead.task_id == task_id)
        count_query = count_query.where(Lead.task_id == task_id)

    if search:
        search_filter = (
            Lead.company_name.ilike(f"%{search}%")
            | Lead.website.ilike(f"%{search}%")
            | Lead.country.ilike(f"%{search}%")
            | Lead.industry.ilike(f"%{search}%")
            | Lead.email.ilike(f"%{search}%")
        )
        query = query.where(search_filter)
        count_query = count_query.where(search_filter)

    if country:
        query = query.where(Lead.country == country)
        count_query = count_query.where(Lead.country == country)

    if industry:
        query = query.where(Lead.industry.ilike(f"%{industry}%"))
        count_query = count_query.where(Lead.industry.ilike(f"%{industry}%"))

    if min_score is not None:
        query = query.where(Lead.match_score >= min_score)
        count_query = count_query.where(Lead.match_score >= min_score)

    # Sort
    sort_columns = {
        "created_at": Lead.created_at,
        "company_name": Lead.company_name,
        "country": Lead.country,
        "industry": Lead.industry,
        "match_score": Lead.match_score,
    }
    sort_col = sort_columns.get(sort_by, Lead.created_at)
    if sort_order == "desc":
        query = query.order_by(sort_col.desc())
    else:
        query = query.order_by(sort_col.asc())

    total = (await db.execute(count_query)).scalar() or 0
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    leads = result.scalars().all()

    task_ids = {lead.task_id for lead in leads}
    task_map: dict[int, Task] = {}
    if task_ids:
        task_result = await db.execute(select(Task).where(Task.id.in_(task_ids)))
        task_map = {task.id: task for task in task_result.scalars().all()}

    items = []
    for lead in leads:
        # Get latest email status for this lead
        email_result = await db.execute(
            select(OutreachEmail)
            .where(OutreachEmail.lead_id == lead.id)
            .order_by(OutreachEmail.created_at.desc())
            .limit(1)
        )
        latest_email = email_result.scalar_one_or_none()
        if latest_email is None:
            email_status = "unwritten"
        elif latest_email.send_status in {
            "sent",
            "delivered",
            "failed",
            "sending",
            "pending",
            "bounced",
            "complained",
        }:
            email_status = latest_email.send_status
        else:
            email_status = "draft"

        items.append(
            LeadResponse(
                id=lead.id,
                source_task_id=lead.task_id,
                source_list=_format_source_list(task_map.get(lead.task_id)),
                user_note=lead.user_note or "",
                company_name=lead.company_name,
                website=lead.website,
                country=lead.country,
                industry=lead.industry,
                company_role=lead.company_role,
                contact_name=clean_contact_name(lead.contact_name),
                email=lead.email,
                phone=lead.phone,
                ai_summary=lead.ai_summary,
                business_match=lead.business_match,
                outreach_suggestion=lead.outreach_suggestion,
                match_score=lead.match_score,
                email_status=email_status,
                email_subject=latest_email.email_subject if latest_email else "",
                email_body=latest_email.email_body if latest_email else "",
            )
        )

    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=(total + page_size - 1) // page_size,
    )


async def update_lead(db: AsyncSession, user_id: int, lead_id: int, req) -> None:
    """Update editable fields for a single lead."""
    from app.models.task import Task
    from fastapi import HTTPException

    result = await db.execute(
        select(Lead).where(
            Lead.id == lead_id,
            Lead.task_id.in_(select(Task.id).where(Task.user_id == user_id)),
        )
    )
    lead = result.scalar_one_or_none()
    if lead is None:
        raise HTTPException(status_code=404, detail="Lead not found")
    if req.contact_name is not None:
        lead.contact_name = req.contact_name
    if req.email is not None:
        lead.email = req.email
    if req.user_note is not None:
        lead.user_note = req.user_note
    await db.commit()


async def delete_leads(db: AsyncSession, user_id: int, lead_ids: list[int]) -> int:
    """Delete leads owned by the current user, including their email records."""
    if not lead_ids:
        return 0

    owned_leads_result = await db.execute(
        select(Lead.id).where(
            Lead.id.in_(lead_ids),
            Lead.task_id.in_(select(Task.id).where(Task.user_id == user_id)),
        )
    )
    owned_lead_ids = [lead_id for (lead_id,) in owned_leads_result.all()]
    if not owned_lead_ids:
        return 0

    await db.execute(delete(OutreachEmail).where(OutreachEmail.lead_id.in_(owned_lead_ids)))
    await db.execute(delete(Lead).where(Lead.id.in_(owned_lead_ids)))
    await db.commit()
    return len(owned_lead_ids)


async def get_leads_with_emails(
    db: AsyncSession,
    user_id: int,
    task_id: int,
    page: int = 1,
    page_size: int = 20,
    search: str | None = None,
) -> PaginatedResponse["LeadEmailResponse"]:
    """Get leads that have emails for a specific email-craft task.

    Uses outreach_emails table to find leads (not Lead.task_id),
    since leads from customer-acquisition have a different task_id.
    """
    from app.schemas.lead import LeadEmailResponse

    # Step 1: Get email records for this task
    email_result = await db.execute(
        select(OutreachEmail).where(OutreachEmail.task_id == task_id)
    )
    email_records = email_result.scalars().all()

    # Step 2: Get unique lead IDs
    lead_ids = list({em.lead_id for em in email_records})
    if not lead_ids:
        return PaginatedResponse(
            items=[], total=0, page=page, page_size=page_size, total_pages=0
        )

    # Build email map: lead_id → OutreachEmail
    email_map = {em.lead_id: em for em in email_records}

    # Step 3: Query leads
    query = select(Lead).where(Lead.id.in_(lead_ids))
    count_query = select(func.count()).select_from(Lead).where(Lead.id.in_(lead_ids))

    if search:
        search_filter = (
            Lead.company_name.ilike(f"%{search}%")
            | Lead.website.ilike(f"%{search}%")
            | Lead.country.ilike(f"%{search}%")
            | Lead.industry.ilike(f"%{search}%")
        )
        query = query.where(search_filter)
        count_query = count_query.where(search_filter)

    total = (await db.execute(count_query)).scalar() or 0

    query = query.order_by(Lead.match_score.desc())
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    leads = result.scalars().all()

    items = []
    for lead in leads:
        em = email_map.get(lead.id)
        items.append(
            LeadEmailResponse(
                id=lead.id,
                company_name=lead.company_name,
                website=lead.website,
                country=lead.country,
                industry=lead.industry,
                company_role=lead.company_role,
                contact_name=clean_contact_name(lead.contact_name),
                email=lead.email,
                phone=lead.phone,
                match_score=lead.match_score,
                email_subject=em.email_subject if em else "",
                email_body=em.email_body if em else "",
            )
        )

    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=(total + page_size - 1) // page_size,
    )


async def export_leads_xlsx(
    db: AsyncSession,
    user_id: int,
    task_id: int | None = None,
    search: str | None = None,
    country: str | None = None,
    industry: str | None = None,
    min_score: float | None = None,
) -> bytes:
    from app.models.task import Task

    task_ids_query = select(Task.id).where(Task.user_id == user_id)
    query = select(Lead).where(Lead.task_id.in_(task_ids_query))

    if task_id is not None:
        query = query.where(Lead.task_id == task_id)
    if search:
        query = query.where(
            Lead.company_name.ilike(f"%{search}%")
            | Lead.website.ilike(f"%{search}%")
            | Lead.country.ilike(f"%{search}%")
            | Lead.industry.ilike(f"%{search}%")
            | Lead.email.ilike(f"%{search}%")
        )
    if country:
        query = query.where(Lead.country == country)
    if industry:
        query = query.where(Lead.industry.ilike(f"%{industry}%"))
    if min_score is not None:
        query = query.where(Lead.match_score >= min_score)

    result = await db.execute(query.order_by(Lead.match_score.desc()))
    leads = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = [
        "Company Name", "Website", "Country", "Industry",
        "Company Role", "Contact Name", "Email", "Phone", "Match Score",
        "AI Summary", "Business Match", "Outreach Suggestion",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="EFF6FF")
    header_font = Font(bold=True, color="1F2937")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for lead in leads:
        ws.append([
            lead.company_name,
            lead.website,
            lead.country,
            lead.industry,
            lead.company_role,
            clean_contact_name(lead.contact_name),
            lead.email,
            lead.phone,
            lead.match_score,
            lead.ai_summary,
            lead.business_match,
            lead.outreach_suggestion,
        ])

    ws.freeze_panes = "A2"
    widths = {
        "A": 28,
        "B": 28,
        "C": 18,
        "D": 22,
        "E": 16,
        "F": 20,
        "G": 28,
        "H": 18,
        "I": 12,
        "J": 52,
        "K": 42,
        "L": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 54

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].bestFit = False

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
