"""Shared CSV pipeline utilities.

Provides a unified CSV schema for the customer-acquisition → email-craft → email-blast
pipeline, replacing Feishu as the data channel when Feishu is unavailable.

All CSV files use utf-8-sig encoding (BOM) for Excel compatibility.
"""

from __future__ import annotations

import csv
import os
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Unified CSV Schema
# ---------------------------------------------------------------------------

# Column definitions: (English key, Chinese header)
CSV_FIELDS = [
    ("company_name", "公司名称"),
    ("website", "网站"),
    ("country", "国家/地区"),
    ("industry", "行业"),
    ("company_role", "公司角色"),
    ("contact_name", "联系人"),
    ("email", "邮箱"),
    ("phone", "电话"),
    ("ai_summary", "AI分析摘要"),
    ("business_match_points", "业务匹配点"),
    ("outreach_content", "开发建议"),
    ("email_sent", "邮件已发送"),
    ("sent_time", "发送时间"),
    ("notes", "备注"),
    ("email_subject", "邮件主题"),
    ("email_draft", "开发信"),
]

ENGLISH_KEYS = [k for k, _ in CSV_FIELDS]
CHINESE_HEADERS = [h for _, h in CSV_FIELDS]

# Reverse lookup: Chinese header → English key
REVERSE_HEADER_MAP = {h: k for k, h in CSV_FIELDS}

# Encoding for all CSV operations
CSV_ENCODING = "utf-8-sig"


def _translate_row(row: dict) -> dict:
    """Translate a CSV row from Chinese headers to English keys.

    Handles both Chinese and English headers, so the CSV works regardless
    of which header style was used.
    """
    result = {}
    for key, value in row.items():
        # Strip whitespace from key
        clean_key = key.strip()
        # Look up in reverse map; if not found, assume it's already an English key
        mapped = REVERSE_HEADER_MAP.get(clean_key, clean_key)
        result[mapped] = value
    return result


# ---------------------------------------------------------------------------
# Read operations
# ---------------------------------------------------------------------------

def read_pipeline_csv(
    filepath: str | Path,
    skip_with_draft: bool = True,
    include_failed: bool = False,
) -> list[dict]:
    """Read pipeline CSV and return records with English keys.

    Each record includes ``_row_index`` (0-based row number, excluding header).

    Args:
        filepath: Path to CSV file.
        skip_with_draft: If True, skip rows that already have an email_draft.
        include_failed: If True, include rows where email_sent is "否" or
            notes contains "发送失败".

    Returns:
        List of dicts with English keys + ``_row_index``.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"CSV file not found: {filepath}")

    records = []
    with open(filepath, "r", encoding=CSV_ENCODING) as f:
        reader = csv.DictReader(f)
        for i, raw_row in enumerate(reader):
            row = _translate_row(raw_row)
            row["_row_index"] = i

            # Filter: skip rows with existing draft
            if skip_with_draft and row.get("email_draft", "").strip():
                continue

            # Filter: skip failed rows unless requested
            if not include_failed:
                sent = row.get("email_sent", "").strip()
                notes = row.get("notes", "").strip()
                if sent == "否" or "发送失败" in notes:
                    continue

            records.append(row)

    return records


def read_all_pipeline_csv(filepath: str | Path) -> list[dict]:
    """Read all rows from pipeline CSV without filtering.

    Used for merge-write-back operations.

    Returns:
        List of dicts with English keys + ``_row_index``.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"CSV file not found: {filepath}")

    records = []
    with open(filepath, "r", encoding=CSV_ENCODING) as f:
        reader = csv.DictReader(f)
        for i, raw_row in enumerate(reader):
            row = _translate_row(raw_row)
            row["_row_index"] = i
            records.append(row)

    return records


# ---------------------------------------------------------------------------
# Write operations
# ---------------------------------------------------------------------------

def write_pipeline_csv(filepath: str | Path, records: list[dict]) -> int:
    """Write records back to pipeline CSV, preserving order by _row_index.

    If the original CSV had columns not in our schema, they are dropped.
    If records have new columns, they are appended to the end.

    Args:
        filepath: Path to CSV file.
        records: List of dicts, each containing ``_row_index``.

    Returns:
        Number of rows written.
    """
    filepath = Path(filepath)

    # Sort by _row_index to preserve original order
    sorted_records = sorted(records, key=lambda r: r.get("_row_index", 0))

    # Determine all column keys present across all records
    all_keys = list(ENGLISH_KEYS)
    for record in sorted_records:
        for key in record:
            if key not in all_keys and key != "_row_index":
                all_keys.append(key)

    # Map to Chinese headers for output
    header_map = {k: h for k, h in CSV_FIELDS}
    # Add any extra keys that aren't in our schema
    for key in all_keys:
        if key not in header_map:
            header_map[key] = key  # Use English key as-is for unknown columns

    headers = [header_map[k] for k in all_keys]

    filepath.parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", newline="", encoding=CSV_ENCODING) as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for record in sorted_records:
            # Build row with Chinese headers
            row = {}
            for eng_key in all_keys:
                cn_header = header_map[eng_key]
                row[cn_header] = record.get(eng_key, "")
            writer.writerow(row)

    return len(sorted_records)


# ---------------------------------------------------------------------------
# Single-row status update
# ---------------------------------------------------------------------------

def update_csv_status(
    filepath: str | Path,
    row_index: int,
    success: bool,
    message_id: str | None = None,
    error: str | None = None,
) -> bool:
    """Update a single row's send status in the pipeline CSV.

    Modifies the CSV in-place: reads all rows, updates the target row,
    writes everything back.

    Args:
        filepath: Path to CSV file.
        row_index: 0-based row index to update.
        success: Whether the email was sent successfully.
        message_id: Resend message ID (if successful).
        error: Error message (if failed).

    Returns:
        True if update succeeded.
    """
    filepath = Path(filepath)
    if not filepath.exists():
        return False

    try:
        all_records = read_all_pipeline_csv(filepath)
    except Exception:
        return False

    # Find and update the target row
    for record in all_records:
        if record.get("_row_index") == row_index:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if success:
                record["email_sent"] = "是"
                record["sent_time"] = now
                if message_id:
                    record["notes"] = f"Resend ID: {message_id}"
            else:
                record["email_sent"] = "否"
                record["sent_time"] = now
                if error:
                    record["notes"] = f"发送失败: {error}"
            break
    else:
        return False

    write_pipeline_csv(filepath, all_records)
    return True


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_excel(csv_path: str | Path, excel_path: str | Path | None = None) -> str:
    """Export pipeline CSV to an Excel (.xlsx) file for sharing.

    Reads the CSV and writes a formatted Excel file with:
    - Chinese headers
    - Auto-adjusted column widths
    - Text wrap for long content columns

    Args:
        csv_path: Path to the source CSV file.
        excel_path: Output Excel path. Defaults to same name with .xlsx extension.

    Returns:
        Path to the created Excel file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")

    csv_path = Path(csv_path)
    if excel_path is None:
        excel_path = csv_path.with_suffix(".xlsx")
    else:
        excel_path = Path(excel_path)

    # Read all records
    all_records = read_all_pipeline_csv(csv_path)

    # Determine columns (use standard schema columns that exist in data)
    col_keys = [k for k in ENGLISH_KEYS if any(r.get(k, "").strip() for r in all_records)]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "客户数据"

    # Write headers (Chinese)
    header_map = {k: h for k, h in CSV_FIELDS}
    for col_idx, key in enumerate(col_keys, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_map.get(key, key))
        cell.font = cell.font.copy(bold=True)

    # Write data rows
    for row_idx, record in enumerate(all_records, 2):
        for col_idx, key in enumerate(col_keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=record.get(key, ""))

    # Long content columns: wrap text, fixed width
    narrow_cols = {"email_draft", "ai_summary", "business_match_points", "outreach_content"}
    for col_idx, key in enumerate(col_keys, 1):
        col_letter = get_column_letter(col_idx)
        if key in narrow_cols:
            ws.column_dimensions[col_letter].width = 50
            for row_idx in range(2, len(all_records) + 2):
                ws.cell(row=row_idx, column=col_idx).alignment = \
                    ws.cell(row=row_idx, column=col_idx).alignment.copy(wrap_text=True)
        else:
            # Auto-fit based on content
            max_len = len(header_map.get(key, key))
            for row_idx in range(2, len(all_records) + 2):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, min(len(val), 50))
            ws.column_dimensions[col_letter].width = max_len + 4

    # Freeze header row
    ws.freeze_panes = "A2"

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(excel_path))
    return str(excel_path)
